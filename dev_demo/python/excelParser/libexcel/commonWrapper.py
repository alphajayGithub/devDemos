#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import os
import sys
import pdb
import logging
import traceback
import re
from pathlib import Path

import tushare as ts
import numpy as np
import pandas as pd
from pandas import DataFrame


import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook

logger = logging.getLogger('commonWrapper')
logger.addHandler(logging.NullHandler())

gStubFlag = False
sheetNamePrefix='output'

dataTitle =   ['万元'
                ,' 万元'
                ,'%'
                ,' %'
                ,'市值'   #  股价*总股本
                ,'Price'  #  股价
                ,'PE'     #  股票软件上的浮动PE
                ,'JPE1'   #  price*总股本/(历史利润+最近季度利润估算年内剩余利润)
                ,'JPE2'
                ,'WPE1'   #  price*总股本/季度利润1*4
                ,'WPE2'
                ,'环比增幅'
                ]

def funcWrapper(function):
    from functools import wraps
    @wraps(function)
    def wrapper(*args,**kwargs):
        functionName = function.__name__
        logger.debug(str(functionName) + " with parameters: " +  str(args))

        logger.info(functionName + ' in ...')
        result = function(*args,**kwargs)
        logger.info(functionName + ' out ...')

        if not result:
            sys.exit(1)

    return wrapper


def parseProfitAndPercent(content):
    matchList = []
    rePatternForProfit ='(?<=净利润)([\.\d-]+)(万|亿)(?:[^\.\d-]+)([\.\d-]+)(万|亿)'
    rePatternForPercent='(?:增长|下降)(?:[^\.\d-]*)([\.\d-]+)(%)(?:[^\.\d-]*)([\.\d-]+)(%)'
    rePattern = [ rePatternForProfit, rePatternForPercent]

    if  content is None or type(content) is not str:
                return []

    for searchPatten in rePattern:
        result = re.search(searchPatten,str(content))
        if  result is not None :
            logger.debug(  result.group()  )
            logger.debug(  result.groups() )
            matchList += result.groups()
        else:
             logger.debug('not found ' + searchPatten + ' in ' + str(content))

    return matchList

def getSaveFile(filenameWithPath):

        filePath, filename =  os.path.split(filenameWithPath)
        callName, extension = os.path.splitext(filename)
        saveFile = filePath+"/" + callName + extension
        saveAsFile = filePath+"/" + callName + str(2) + extension
        logger.debug( 'current absolute dir:' + os.path.dirname(os.path.abspath(filenameWithPath)) )
        logger.debug( 'filenameWithPath:' + filenameWithPath)
        logger.debug( 'filePath|name|extension:' + filePath + " " + callName + extension)
        logger.debug( 'Save into the file: ' + saveAsFile)

        return saveAsFile

def initHeaderAndformatCode(srcSheet, myWorkSheet):
    logger.debug("max_row : "    + str(srcSheet.max_row))
    logger.debug("max_column : " + str(srcSheet.max_column))


    targetContentColumn=0 
    for cIndex in range(1,srcSheet.max_column):
        if "代码" in str(srcSheet.cell(1,cIndex).value):
            targetContentColumn = cIndex

    #code = srcSheet.cell(5,targetContentColumn).value   
    #print(  len(str(code)) )   
    #print(  str(code).zfill(6) )

    for r in range(1, srcSheet.max_row+1):
            value = srcSheet.cell(r,targetContentColumn).value 
            if type(value) is int :  # and len(str(value)) < 6
                 #logger.info( "value is " + str(value))
                 myWorkSheet.cell(r,targetContentColumn).value = str(value).zfill(6)
                 #logger.info(myWorkSheet.cell(r,targetContentColumn).value )

    #cellRange = srcSheet['F2':'F4']

    myWorkSheet.insert_cols(0, len(dataTitle))
    dataOffset =0  
    for index in range(0,len(dataTitle)):
        column = (index+1) + dataOffset
        myWorkSheet.cell(1,column).value = dataTitle[index]



def writeDataForYuGao(srcSheet, myWorkSheet):
    offset =0  
    
    for cIndex in range(1,srcSheet.max_column):
        if "预告摘要" in str(srcSheet.cell(1,cIndex).value):
            targetContentColumn = cIndex
            logger.info("Found content column index " + str(targetContentColumn))

    for row in range(2, srcSheet.max_row+1):
        content = srcSheet.cell(row,targetContentColumn).value
        matchList = parseProfitAndPercent(content)
        logger.info(matchList)
        for index in range(0,len(matchList)):
            writeColumn = int(index/2+ 1) + offset
            if index % 2 ==0:
                myWorkSheet.cell(row, writeColumn).value=matchList[index]



def  setTSdata(srcData, myWorkSheet ):
    targetContentColumn=0 
    for cIndex in range(1,myWorkSheet.max_column):
        if "代码" in str(myWorkSheet.cell(1,cIndex).value):
            targetContentColumn = cIndex
    
    print(targetContentColumn)

    profitIndex = dataTitle.index('万元')+1
    profitIndex2 = profitIndex+1
    WPEIndex = dataTitle.index('WPE1')+1
    WPEIndex2 = dataTitle.index('WPE2')+1

    priceColumnIndex = dataTitle.index('Price')+1
    mktcapIndex = dataTitle.index('市值')+1
    perIndex = dataTitle.index('PE')+1

    #logger.info(dataTitle.index('市值'))
    #logger.info(dataTitle.index('Price'))
    
    #logger.critical(srcData['code'])
    for r in range(1, myWorkSheet.max_row+1):
            key = myWorkSheet.cell(r,targetContentColumn).value 
            #print(type(value))
            #print((value))
            #print(len(key))
            if  (type(key) is str and  len(key) == 6 ):
                 logger.info( "code is " + key)

                

                 #myWorkSheet.cell(r,targetContentColumn).value = str(value).zfill(6)
                 #logger.info(myWorkSheet.cell(r,targetContentColumn).value )
                 tradeValue  = srcData[ srcData['code'] == int(key) ]['trade'].values
                 mktcapValue = srcData[ srcData['code'] == int(key) ]['mktcap'].values
                 perValue    = srcData[ srcData['code'] == int(key) ]['per'].values

                 if tradeValue.size == 0 :
                       logger.critical("No key in trade data: " +  str(key))
                       continue

                 #print(type(str(tradeValue[0])))
                 #print (float(tradeValue))
                 myWorkSheet.cell(r,priceColumnIndex).value =  str(tradeValue[0])
                 myWorkSheet.cell(r,mktcapIndex).value      =  str(mktcapValue[0])
                 myWorkSheet.cell(r,perIndex).value         =  str(perValue[0])
                 #print(myWorkSheet.cell(r,priceColumnIndex).value)

                 if myWorkSheet.cell(r,profitIndex).value is not None:
                    profitValue = myWorkSheet.cell(r,profitIndex).value
                    profitValue2 = myWorkSheet.cell(r,profitIndex2).value    
                    myWorkSheet.cell(r,WPEIndex).value  =  float(mktcapValue[0])/(float(profitValue)*4)
                    myWorkSheet.cell(r,WPEIndex2).value  = float(mktcapValue[0])/(float(profitValue2)*4)
    
    '''
    #print( srcData[ srcData['code'] ==688399 ]['trade'].values)
    for dataIndex  in srcData.index.values:
        #print( srcData['code'][dataIndex])
        if srcData['code'][dataIndex] == 688399:
                    logger.info( srcData['trade'][dataIndex]  )
    '''

def createMySheet(workBook, srcSheet, desSheetName):
    if  desSheetName in workBook.get_sheet_names():
                    workBook.remove_sheet(workBook.get_sheet_by_name(desSheetName))

    #myWorkSheet = workBook.create_sheet(desSheetName)
    
    myWorkSheet = workBook.copy_worksheet(srcSheet) 
    myWorkSheet.title = desSheetName
    return myWorkSheet

@funcWrapper
def readExcel(filenameWithPath):
    logger.info("read file: " + filenameWithPath)
    srcFile = Path(filenameWithPath)

    if not srcFile.exists():
        logger.critical("file not exist")
        return
    else:
        #https://www.jianshu.com/p/576c0c6fa3d9

        workBook = load_workbook(filenameWithPath)
        logger.info(workBook.get_sheet_names())
        #current_ws = workBook.active
        #sheets = workBook.worksheets
        #srcSheet = sheets[0]

        
        filename = 'output/data.xls'
        if not os.path.exists(filename):
            currentDayData = ts.get_today_all()
            logger.debug(currentDayData)

            #currentDayData.drop(currentDayData[currentDayData['turnoverratio']==0].index,inplace=True)
            currentDayData.to_excel(filename,sheet_name='All')
        else:
            currentDayData=pd.read_excel(filename,sheet_name='All')

           

        for parseSheetName in workBook.get_sheet_names():
            logger.info( "current Sheet is:  " + parseSheetName)
            srcSheet =  workBook.get_sheet_by_name(parseSheetName)



            if "预告" in parseSheetName:                   
                    logger.debug("预告: " + parseSheetName)
                    myWorkSheetName =  sheetNamePrefix + "_" + parseSheetName
                    myWorkSheet = createMySheet(workBook,srcSheet, myWorkSheetName)

                    initHeaderAndformatCode(srcSheet,myWorkSheet)
                    writeDataForYuGao(srcSheet,myWorkSheet)
                    setTSdata(currentDayData,myWorkSheet)

            elif "快报" in parseSheetName or "公告" in parseSheetName:
                    logger.debug("快报: " + parseSheetName)

                    myWorkSheetName =  sheetNamePrefix + "_" + parseSheetName
                    myWorkSheet = createMySheet(workBook,srcSheet, myWorkSheetName)

                    initHeaderAndformatCode(srcSheet,myWorkSheet)
                    #writeDataForGongGao(srcSheet,myWorkSheet)
                    setTSdata(currentDayData,myWorkSheet)

            else :
                    logger.error("unexpect sheet")

        workBook.save(getSaveFile(filenameWithPath))
        workBook.close()


@funcWrapper
def create_sheet_cell(nSheet, nCell, filename):
        wb = Workbook()

        for n in range(nSheet):
            wb.create_sheet(index=n, title="Sheet_" + str(n+1))

        currentSheet = wb.active
        for r in range(1, nCell+1):
            for c in range(1, nCell+1):
                 cell = currentSheet.cell(row=r, column=c, value= r * c)

        wb.save(saveFile= "output/" + str(filename))


def create_workbook():
        nSheet = int(input("输入电子表格的个数（整数）："))
        nCell =  int(input("输入乘法表的最大值（整数）："))
        saveFileName =  str(input("最终存储的文件名："))
        create_sheet_cell(nSheet, nCell, saveFileName)


def  debug():
        sys.exit(0)